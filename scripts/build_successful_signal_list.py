#!/usr/bin/env python3
"""Build successful_signal_list.xlsx — sinyal dengan kenaikan > 10%.

Membaca semua data/performance/daily/{scalping,swing}_*.csv (sinyal yang sudah
evaluated), filter ke baris dengan kenaikan High ATAU kenaikan Close > 10%,
lalu lookup harga Open dari data/raw/{ticker}.parquet pada tanggal evaluasi.

Output: data/performance/successful_signal_list.xlsx

Usage:
    python scripts/build_successful_signal_list.py
"""
import sys
from pathlib import Path

repo_root = Path(__file__).parent.parent
sys.path.insert(0, str(repo_root))

import pandas as pd
from loguru import logger

_DAILY_DIR = repo_root / "data" / "performance" / "daily"
_RAW_DIR = repo_root / "data" / "raw"
_BUNDLE_PATH = repo_root / "data" / "published" / "ohlc_recent.parquet"
_OUT_PATH = repo_root / "data" / "performance" / "successful_signal_list.xlsx"
_MIN_GAIN_PCT = 10.0

# Per-ticker raw OHLC cache so each parquet is read at most once.
_raw_cache: dict[str, pd.DataFrame] = {}
_bundle_cache: pd.DataFrame | None = None


def _load_all_daily() -> pd.DataFrame:
    frames = []
    for f in sorted(_DAILY_DIR.glob("*.csv")):
        if not (f.name.startswith("scalping_") or f.name.startswith("swing_")):
            continue
        try:
            csv_df = pd.read_csv(f)
            if not csv_df.empty:
                frames.append(csv_df)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Skip {}: {}", f.name, exc)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def _load_bundle() -> pd.DataFrame:
    global _bundle_cache
    if _bundle_cache is None:
        if _BUNDLE_PATH.exists():
            df = pd.read_parquet(_BUNDLE_PATH)
            df["date"] = pd.to_datetime(df["date"]).dt.strftime("%Y-%m-%d")
            _bundle_cache = df
        else:
            _bundle_cache = pd.DataFrame()
    return _bundle_cache


def _get_open(ticker: str, eval_date: str) -> float | None:
    # 1) per-ticker raw parquet (most complete local history)
    if ticker not in _raw_cache:
        path = _RAW_DIR / f"{ticker}.parquet"
        if not path.exists():
            _raw_cache[ticker] = pd.DataFrame()
        else:
            df = pd.read_parquet(path)
            df["date"] = pd.to_datetime(df["date"]).dt.strftime("%Y-%m-%d")
            _raw_cache[ticker] = df
    df = _raw_cache[ticker]
    if not df.empty:
        row = df[df["date"] == eval_date]
        if not row.empty:
            return float(row.iloc[0]["open"])

    # 2) committed recent-OHLC bundle (covers gaps in the local raw cache)
    bundle = _load_bundle()
    if not bundle.empty:
        row = bundle[(bundle["ticker"] == ticker) & (bundle["date"] == eval_date)]
        if not row.empty:
            return float(row.iloc[0]["open"])

    return None


def main() -> None:
    df = _load_all_daily()
    if df.empty:
        logger.warning("Tidak ada data di {} — keluar.", _DAILY_DIR)
        return

    df = df[df["status"] == "evaluated"].copy()
    df = df[(df["pct_high"] > _MIN_GAIN_PCT) | (df["pct_close"] > _MIN_GAIN_PCT)]
    df = df.drop_duplicates(subset=["ticker", "eval_date", "high", "close"])
    df = df.sort_values("pct_close", ascending=False).reset_index(drop=True)

    df["open"] = [
        _get_open(r["ticker"], r["eval_date"]) for _, r in df.iterrows()
    ]

    out = pd.DataFrame({
        "Nama Saham": df["ticker"],
        "Tanggal": df["eval_date"],
        "Kenaikan High (%)": df["pct_high"],
        "Kenaikan Close (%)": df["pct_close"],
        "Open": df["open"],
        "High": df["high"],
        "Close": df["close"],
    })

    _write_excel(out)
    logger.info("Selesai: {} sinyal sukses (>{:.0f}%) -> {}", len(out), _MIN_GAIN_PCT, _OUT_PATH)


def _write_excel(df: pd.DataFrame) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception as exc:  # noqa: BLE001
        logger.warning("openpyxl tidak tersedia — fallback ke CSV ({}).", exc)
        df.to_csv(_OUT_PATH.with_suffix(".csv"), index=False)
        return

    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(color="FFFFFF", bold=True)

    _OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(_OUT_PATH, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="Successful Signals", index=False, startrow=0)
        ws = xw.book["Successful Signals"]
        for j, _ in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=j)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for i in range(2, len(df) + 2):
            ws.cell(row=i, column=3).number_format = '0.00"%"'
            ws.cell(row=i, column=4).number_format = '0.00"%"'
            for col in (5, 6, 7):
                ws.cell(row=i, column=col).number_format = "#,##0"
        for j, w in enumerate([14, 12, 18, 19, 10, 10, 10], 1):
            ws.column_dimensions[chr(64 + j)].width = w


if __name__ == "__main__":
    main()
