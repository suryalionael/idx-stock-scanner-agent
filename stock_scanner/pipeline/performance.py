"""Signal List performance tracking — daily archive + next-session evaluation.

A reporting/tracking layer on top of the screener (the screener logic is NOT
changed). For each trading day it:

  1. Archives the Swing and Scalping signal lists (permanent snapshot).
  2. Evaluates each signal against the NEXT valid market session's OHLC.
  3. Writes daily CSV + a ready-to-read Excel workbook, and updates the
     permanent archive / results CSVs.

Convention (per product spec)
-----------------------------
  Signal date     = the session the ticker appeared in the Swing/Scalping list
  Evaluation date = next valid trading session after the signal date
  Entry  (Open)   = evaluation-date open
  High            = evaluation-date high
  Close           = evaluation-date close
  Percentage High = (High  - Open) / Open * 100
  Percentage Close= (Close - Open) / Open * 100
  W/L             = "W" if Close > Open else "L"
  Win Rate        = #W / #evaluated * 100

If the next session's OHLC is not available yet, the record is kept as
status="pending" (NOT faked) and filled on the next run.

Strategy membership
  Swing    : signal ∈ {BREAKOUT, PRE_MARKUP}
  Scalping : scalping_label ∈ {SCALPING_HIGH}

Outputs (all under data/performance/)
  signals_archive.csv                 — permanent signal snapshots (append/dedup)
  signal_results.csv                  — permanent evaluated results (upsert)
  daily/swing_YYYY-MM-DD.csv           — Swing signals for that signal date
  daily/scalping_YYYY-MM-DD.csv        — Scalping signals for that signal date
  daily/signal_list_YYYY-MM-DD.xlsx    — Swing + Scalping sheets + win-rate
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from loguru import logger

_ROOT = Path(__file__).parent.parent.parent
_SIGNALS_DIR = _ROOT / "data" / "signals"
_RAW_DIR = _ROOT / "data" / "raw"
_PERF_DIR = _ROOT / "data" / "performance"
_DAILY_DIR = _PERF_DIR / "daily"

_ARCHIVE_CSV = _PERF_DIR / "signals_archive.csv"
_RESULTS_CSV = _PERF_DIR / "signal_results.csv"

SWING_SIGNALS = {"BREAKOUT", "PRE_MARKUP"}
SCALPING_LABELS = {"SCALPING_HIGH"}

# Column order for the daily table. Reference price = PREVIOUS CLOSE (the close
# of the signal session); High/Close are the next session's, measured vs Prev.
_RESULT_COLS = [
    "signal_date", "eval_date", "strategy", "ticker", "signal",
    "prev", "close", "high", "pct_high", "pct_close", "wl", "status",
]


# ---------------------------------------------------------------------------
# Signal-list extraction
# ---------------------------------------------------------------------------

def _signal_list(df: pd.DataFrame, strategy: str) -> pd.DataFrame:
    """Return [ticker, signal] for one strategy from a signals frame."""
    if df is None or df.empty or "ticker" not in df.columns:
        return pd.DataFrame(columns=["ticker", "signal"])
    if strategy == "swing":
        if "signal" not in df.columns:
            return pd.DataFrame(columns=["ticker", "signal"])
        sub = df[df["signal"].astype(str).isin(SWING_SIGNALS)]
        out = sub[["ticker", "signal"]].copy()
    else:  # scalping
        if "scalping_label" not in df.columns:
            return pd.DataFrame(columns=["ticker", "signal"])
        sub = df[df["scalping_label"].astype(str).isin(SCALPING_LABELS)]
        out = sub[["ticker"]].copy()
        out["signal"] = sub["scalping_label"].astype(str).values
    return out.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Next-session OHLC lookup (never fabricates)
# ---------------------------------------------------------------------------

def _next_session_ohlc(ticker: str, signal_date: str, raw_dir: Path) -> Optional[dict]:
    """Previous-close reference + next valid session's high/close for `ticker`.

    Returns {prev_close, eval_date, high, close} or None when not available yet
    (caller marks the record pending). `prev_close` is the close of the session
    immediately BEFORE the evaluation session (i.e. the signal-session close).
    Skips zero-volume / NaN bars.
    """
    path = raw_dir / f"{ticker}.parquet"
    if not path.exists():
        return None
    try:
        df = pd.read_parquet(path, columns=["date", "open", "high", "low", "close", "volume"])
    except Exception:  # noqa: BLE001
        try:
            df = pd.read_parquet(path)
        except Exception:  # noqa: BLE001
            return None
    if df.empty or "date" not in df.columns:
        return None
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.tz_localize(None).dt.normalize()
    df = df.sort_values("date").reset_index(drop=True)
    if "volume" in df.columns:
        df = df[pd.to_numeric(df["volume"], errors="coerce").fillna(0) > 0].reset_index(drop=True)
    sig_ts = pd.Timestamp(signal_date)
    fwd_idx = df.index[df["date"] > sig_ts]
    for i in fwd_idx:
        if i == 0:  # no prior session → no previous close
            continue
        ev = df.loc[i]
        prev_close = df.loc[i - 1, "close"]
        h, c = ev.get("high"), ev.get("close")
        if (pd.notna(prev_close) and float(prev_close) > 0
                and pd.notna(h) and pd.notna(c)):
            return {
                "prev_close": float(prev_close),
                "eval_date": ev["date"].strftime("%Y-%m-%d"),
                "high": float(h), "close": float(c),
            }
    return None


def _evaluate_row(signal_date: str, strategy: str, ticker: str, signal: str,
                  raw_dir: Path) -> dict:
    """Build one result row — evaluated or pending. Reference = previous close."""
    base = {"signal_date": signal_date, "strategy": strategy, "ticker": ticker,
            "signal": signal, "eval_date": None, "prev": None, "close": None,
            "high": None, "pct_high": None, "pct_close": None, "wl": None,
            "status": "pending"}
    ohlc = _next_session_ohlc(ticker, signal_date, raw_dir)
    if ohlc is None:
        return base
    p, h, c = ohlc["prev_close"], ohlc["high"], ohlc["close"]
    base.update({
        "eval_date": ohlc["eval_date"],
        "prev": round(p, 2), "high": round(h, 2), "close": round(c, 2),
        "pct_high": round((h - p) / p * 100, 2),
        "pct_close": round((c - p) / p * 100, 2),
        "wl": "W" if c > p else "L",
        "status": "evaluated",
    })
    return base


# ---------------------------------------------------------------------------
# Persistent archive / results helpers (idempotent upserts)
# ---------------------------------------------------------------------------

def _read_csv(path: Path) -> pd.DataFrame:
    if path.exists():
        try:
            return pd.read_csv(path)
        except Exception:  # noqa: BLE001
            return pd.DataFrame()
    return pd.DataFrame()


def _upsert(existing: pd.DataFrame, new: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    if existing is None or existing.empty:
        return new.reset_index(drop=True)
    if new is None or new.empty:
        return existing.reset_index(drop=True)
    for k in keys:
        if k not in existing.columns:
            existing[k] = None
    merged = pd.concat([existing, new], ignore_index=True)
    merged = merged.drop_duplicates(subset=keys, keep="last").reset_index(drop=True)
    return merged


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _write_excel(path: Path, swing: pd.DataFrame, scalping: pd.DataFrame, signal_date: str) -> None:
    """Ready-to-read workbook: a sheet per strategy, % formatting, W/L colour,
    win-rate summary at the top. Best-effort (skipped if openpyxl missing)."""
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:  # noqa: BLE001
        logger.warning("openpyxl not available — skipping Excel ({}).", exc)
        return

    disp_cols = ["ticker", "signal", "prev", "close", "high",
                 "pct_high", "pct_close", "wl", "eval_date", "status"]
    headers = ["Signal", "Type", "Prev", "Close", "High",
               "Percentage High", "Percentage Close", "W/L", "Eval Date", "Status"]

    win_fill = PatternFill("solid", fgColor="C6EFCE")
    loss_fill = PatternFill("solid", fgColor="FFC7CE")
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(color="FFFFFF", bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for sheet, df in (("Swing", swing), ("Scalping", scalping)):
            ev = df[df["status"] == "evaluated"] if not df.empty else df
            n = len(ev)
            wins = int((ev["wl"] == "W").sum()) if n else 0
            wr = round(wins / n * 100, 1) if n else 0.0
            pend = int((df["status"] == "pending").sum()) if not df.empty else 0

            ws = xw.book.create_sheet(sheet)
            ws["A1"] = f"{sheet} — Review sesi market {signal_date}"
            ws["A1"].font = Font(bold=True, size=13)
            ws["A2"] = (f"Win Rate: {wr}%  ({wins}W / {n - wins}L of {n} evaluated"
                        + (f", {pend} pending)" if pend else ")"))
            ws["A2"].font = Font(bold=True, color="2563EB")

            start = 4
            for j, h in enumerate(headers, 1):
                cell = ws.cell(row=start, column=j, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")

            if df.empty:
                ws.cell(row=start + 1, column=1, value="(tidak ada sinyal)")
            else:
                for i, (_, r) in enumerate(df.iterrows(), start=start + 1):
                    for j, col in enumerate(disp_cols, 1):
                        ws.cell(row=i, column=j, value=r.get(col))
                    # % formatting
                    for jc in (6, 7):  # pct_high, pct_close columns
                        ws.cell(row=i, column=jc).number_format = '0.00"%"'
                    # W/L colour
                    wl = r.get("wl")
                    if wl == "W":
                        ws.cell(row=i, column=8).fill = win_fill
                    elif wl == "L":
                        ws.cell(row=i, column=8).fill = loss_fill
            # widths
            for j, w in enumerate([12, 13, 10, 10, 10, 15, 16, 6, 12, 11], 1):
                ws.column_dimensions[chr(64 + j)].width = w

        # openpyxl always creates a default "Sheet" — remove it.
        if "Sheet" in xw.book.sheetnames:
            del xw.book["Sheet"]


# ---------------------------------------------------------------------------
# Per-date processing + orchestrator
# ---------------------------------------------------------------------------

def process_date(signal_date: str, signals_dir: Path | None = None,
                 raw_dir: Path | None = None, perf_dir: Path | None = None,
                 exclude_suspended: bool = True) -> dict:
    """Archive + evaluate one signal date. Returns per-strategy summary."""
    signals_dir = signals_dir or _SIGNALS_DIR
    raw_dir = raw_dir or _RAW_DIR
    perf_dir = perf_dir or _PERF_DIR
    daily_dir = perf_dir / "daily"

    f = signals_dir / f"{signal_date}.parquet"
    if not f.exists():
        f = signals_dir / f"{signal_date}.csv"
    if not f.exists():
        return {}
    df = pd.read_parquet(f) if f.suffix == ".parquet" else pd.read_csv(f)

    if exclude_suspended:
        try:
            from stock_scanner.pipeline.suspension import filter_active
            df = filter_active(df)
        except Exception:  # noqa: BLE001
            pass

    archive_rows, result_rows, per_strategy = [], [], {}
    daily_dir.mkdir(parents=True, exist_ok=True)

    for strategy in ("swing", "scalping"):
        lst = _signal_list(df, strategy)
        rows = [
            _evaluate_row(signal_date, strategy, str(r["ticker"]), str(r["signal"]), raw_dir)
            for _, r in lst.iterrows()
        ]
        sdf = pd.DataFrame(rows, columns=_RESULT_COLS) if rows else pd.DataFrame(columns=_RESULT_COLS)

        # NOTE: daily review files are written keyed by EVAL DATE (the market
        # session reviewed) in _rebuild_daily_reviews(), not by signal_date —
        # see run_performance(). This is what makes the review dated by today's
        # market session instead of lagging one day behind on the signal date.

        for _, r in lst.iterrows():
            archive_rows.append({"signal_date": signal_date, "strategy": strategy,
                                 "ticker": str(r["ticker"]), "signal": str(r["signal"])})
        result_rows.extend(rows)

        ev = sdf[sdf["status"] == "evaluated"]
        n, wins = len(ev), int((ev["wl"] == "W").sum()) if len(ev) else 0
        per_strategy[strategy] = {
            "signals": len(sdf), "evaluated": n, "wins": wins, "losses": n - wins,
            "pending": int((sdf["status"] == "pending").sum()),
            "win_rate": round(wins / n * 100, 1) if n else None,
            "df": sdf,
        }

    # Permanent stores (upsert)
    _ARCHIVE_CSV.parent.mkdir(parents=True, exist_ok=True)
    _upsert(_read_csv(_ARCHIVE_CSV), pd.DataFrame(archive_rows),
            ["signal_date", "strategy", "ticker"]).to_csv(_ARCHIVE_CSV, index=False)
    _upsert(_read_csv(_RESULTS_CSV), pd.DataFrame(result_rows, columns=_RESULT_COLS),
            ["signal_date", "strategy", "ticker"]).to_csv(_RESULTS_CSV, index=False)

    return per_strategy


def _rebuild_daily_reviews(perf_dir: Path | None = None) -> list[str]:
    """(Re)write daily review files keyed by EVAL DATE — the market session being
    reviewed — from the evaluated rows in signal_results.csv. This is the date the
    user sees: a session that closed today (06-10) produces *_2026-06-10 files,
    not a file dated by the prior signal session (06-09).

    Returns the list of eval dates written.
    """
    perf_dir = perf_dir or _PERF_DIR
    daily_dir = perf_dir / "daily"
    daily_dir.mkdir(parents=True, exist_ok=True)

    res = _read_csv(perf_dir / "signal_results.csv")
    if res.empty or "eval_date" not in res.columns:
        return []
    ev = res[(res["status"] == "evaluated") & res["eval_date"].notna()].copy()
    if ev.empty:
        return []

    written: list[str] = []
    for eval_date, grp in ev.groupby(ev["eval_date"].astype(str)):
        sw = grp[grp["strategy"] == "swing"].reset_index(drop=True)
        sc = grp[grp["strategy"] == "scalping"].reset_index(drop=True)
        sw[_RESULT_COLS].to_csv(daily_dir / f"swing_{eval_date}.csv", index=False)
        sc[_RESULT_COLS].to_csv(daily_dir / f"scalping_{eval_date}.csv", index=False)
        _write_excel(daily_dir / f"signal_list_{eval_date}.xlsx", sw, sc, eval_date)
        written.append(eval_date)
    return sorted(written)


def _reevaluate_pending(raw_dir: Path | None = None, perf_dir: Path | None = None) -> int:
    """Fill still-pending results directly from signal_results.csv + raw OHLC.

    Every pending row already carries (signal_date, strategy, ticker, signal), so
    a signal dated X is evaluated the moment X+1's bar lands in raw_dir — WITHOUT
    needing the original per-date signal file. That matters because data/signals/
    is gitignored and absent on a fresh CI checkout, which previously left old
    pending rows orphaned (eval_date froze at the last fully-evaluated session and
    review dates like 2026-06-11 / 06-12 were skipped). Running this every pass
    guarantees eval_date advances as soon as the next session's data exists.

    Returns the number of rows newly evaluated this pass.
    """
    raw_dir = raw_dir or _RAW_DIR
    perf_dir = perf_dir or _PERF_DIR
    res = _read_csv(perf_dir / "signal_results.csv")
    if res.empty or "status" not in res.columns:
        return 0
    pend = res[res["status"].astype(str) == "pending"].copy()
    if pend.empty:
        return 0
    rows = [
        _evaluate_row(str(r["signal_date"]), str(r["strategy"]),
                      str(r["ticker"]), str(r["signal"]), raw_dir)
        for _, r in pend.iterrows()
    ]
    upd = pd.DataFrame(rows, columns=_RESULT_COLS)
    merged = _upsert(res, upd, ["signal_date", "strategy", "ticker"])
    merged.to_csv(perf_dir / "signal_results.csv", index=False)
    return int((upd["status"] == "evaluated").sum())


def run_performance(signals_dir: Path | None = None, raw_dir: Path | None = None,
                    perf_dir: Path | None = None, dates: list[str] | None = None) -> dict:
    """Archive + (re)evaluate. Default: every available signal date (so pending
    rows get filled once their next session exists). Returns {date: summary}."""
    signals_dir = signals_dir or _SIGNALS_DIR
    raw_dir = raw_dir or _RAW_DIR
    perf_dir = perf_dir or _PERF_DIR
    if dates is None:
        dates = sorted(p.stem for p in signals_dir.glob("*.parquet"))
        dates += [p.stem for p in signals_dir.glob("*.csv") if p.stem not in dates]
        dates = sorted(set(dates))
    out = {}
    for d in dates:
        try:
            res = process_date(d, signals_dir, raw_dir, perf_dir)
            if res:
                out[d] = res
        except Exception as exc:  # noqa: BLE001
            logger.warning("performance: failed on {}: {}", d, exc)

    # Fill any still-pending rows straight from the committed results.csv, so a
    # missing/ephemeral signal file can never freeze eval_date (the 06-11/06-12
    # gap). This is the authoritative step that advances the review date.
    filled = _reevaluate_pending(raw_dir, perf_dir)
    if filled:
        logger.info("Performance: filled {} pending signal(s) from refreshed OHLC.", filled)

    # Daily review files keyed by eval (market) date — the date the user reviews.
    review_dates = _rebuild_daily_reviews(perf_dir)
    if review_dates:
        logger.info("Performance reviews written thru market date {} ({} sessions).",
                    review_dates[-1], len(review_dates))
    return out


# ---------------------------------------------------------------------------
# Read helpers for dashboard / Telegram
# ---------------------------------------------------------------------------

def load_results(perf_dir: Path | None = None) -> pd.DataFrame:
    perf_dir = perf_dir or _PERF_DIR
    return _read_csv(perf_dir / "signal_results.csv")


# ---------------------------------------------------------------------------
# Evening (18:00 WIB) entrypoint — fetch same-day OHLC, then evaluate
# ---------------------------------------------------------------------------

def _pending_tickers(perf_dir: Path | None = None) -> set[str]:
    df = load_results(perf_dir)
    if df.empty or "status" not in df.columns:
        return set()
    return set(df.loc[df["status"] == "pending", "ticker"].astype(str))


def _refresh_raw_for(tickers: list[str], raw_dir: Path, lookback_years: int = 1) -> int:
    """Fetch the latest OHLC for specific tickers, INCLUDING today's just-closed
    session (end = today+1, since yfinance's end is exclusive and the standard
    incremental_update uses end=today which omits the current day). Run after
    market close so today's bar is final. Reuses fetch_yfinance helpers; does not
    touch the screener. Returns count attempted. Best-effort per ticker.
    """
    if not tickers:
        return 0
    from datetime import datetime, timedelta

    from stock_scanner.pipeline.fetch_yfinance import (
        YFinanceFetcher,
        default_start_date,
        load_raw,
        save_raw,
    )
    fetcher = YFinanceFetcher(batch_size=20)
    end = (datetime.today() + timedelta(days=1)).strftime("%Y-%m-%d")  # inclusive of today
    n = 0
    for t in tickers:
        try:
            existing = load_raw(t, raw_dir)
            if existing.empty:
                start = default_start_date(lookback_years)
            else:
                last = pd.to_datetime(existing["date"]).max()
                start = (last + timedelta(days=1)).strftime("%Y-%m-%d")
            if start >= end:
                n += 1
                continue
            new = fetcher.fetch_single(t, start, end)
            if new is None or new.empty:
                n += 1
                continue
            combined = pd.concat([existing, new], ignore_index=True)
            combined["date"] = pd.to_datetime(combined["date"]).dt.tz_localize(None).dt.normalize()
            combined = (combined.drop_duplicates(subset=["date", "ticker"])
                        .sort_values("date").reset_index(drop=True))
            save_raw(t, combined, raw_dir)
            n += 1
        except Exception as exc:  # noqa: BLE001
            logger.debug("refresh_raw {}: {}", t, exc)
    return n


def run_performance_evening(signals_dir: Path | None = None, raw_dir: Path | None = None,
                            perf_dir: Path | None = None, lookback_years: int = 1) -> dict:
    """18:00 WIB entrypoint. Refreshes OHLC for the tickers awaiting evaluation
    (pending results + the latest signal date's lists), so that if the market
    data source has already published today's close, pending rows are filled on
    THIS run instead of waiting for the next morning. Then runs the standard
    performance pass. Does NOT touch the screener.
    """
    signals_dir = signals_dir or _SIGNALS_DIR
    raw_dir = raw_dir or _RAW_DIR
    perf_dir = perf_dir or _PERF_DIR

    need = _pending_tickers(perf_dir)
    files = sorted(signals_dir.glob("*.parquet"))
    if files:
        try:
            df = pd.read_parquet(files[-1])
            for strat in ("swing", "scalping"):
                need |= set(_signal_list(df, strat)["ticker"].astype(str))
        except Exception:  # noqa: BLE001
            pass

    refreshed = _refresh_raw_for(sorted(need), raw_dir, lookback_years)
    logger.info("Evening performance: refreshed OHLC for {} ticker(s) awaiting eval.", refreshed)
    return run_performance(signals_dir, raw_dir, perf_dir)


def win_rate_recap(eval_date: str, perf_dir: Path | None = None) -> dict:
    """Per-strategy win-rate summary for one EVAL (market) date — the reviewed
    session — for Telegram/dashboard."""
    df = load_results(perf_dir)
    out = {}
    if df.empty:
        return out
    for strat in ("swing", "scalping"):
        sub = df[(df["eval_date"].astype(str) == str(eval_date)) & (df["strategy"] == strat)
                 & (df["status"] == "evaluated")]
        n = len(sub)
        wins = int((sub["wl"] == "W").sum()) if n else 0
        out[strat] = {"signals": n, "wins": wins, "losses": n - wins,
                      "win_rate": round(wins / n * 100, 1) if n else None}
    return out


def latest_evaluated_date(perf_dir: Path | None = None) -> Optional[str]:
    """Most recent reviewed market session (max eval_date among evaluated rows)."""
    df = load_results(perf_dir)
    if df.empty or "eval_date" not in df.columns:
        return None
    ev = df[(df["status"] == "evaluated") & df["eval_date"].notna()]
    return None if ev.empty else str(ev["eval_date"].max())


def send_daily_excel_telegram(eval_date: str | None = None,
                              perf_dir: Path | None = None) -> bool:
    """Best-effort: upload the daily Signal List performance .xlsx to Telegram with
    a short win-rate caption. Fail-safe — never raises and never blocks file
    generation. Token/chat come from TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID (repo
    secrets). Defaults to the latest evaluated session. Returns True on success.
    """
    perf_dir = perf_dir or _PERF_DIR
    try:
        eval_date = eval_date or latest_evaluated_date(perf_dir)
        if not eval_date:
            logger.info("Telegram: no evaluated session yet — nothing to send.")
            return False
        xlsx = perf_dir / "daily" / f"signal_list_{eval_date}.xlsx"
        if not xlsx.exists():
            logger.warning("Telegram: {} not found — skip.", xlsx.name)
            return False

        recap = win_rate_recap(eval_date, perf_dir)

        def _line(strat: str) -> str:
            d = recap.get(strat, {})
            wr = d.get("win_rate")
            tail = f" · WR {wr}%" if wr is not None else ""
            return f"{strat.title()}: {d.get('wins', 0)}W / {d.get('signals', 0)}{tail}"

        caption = ("📋 Signal List Performance — sesi market "
                   f"{eval_date}\n{_line('swing')}\n{_line('scalping')}")
        from stock_scanner.alerts.telegram_alert import TelegramAlert
        res = TelegramAlert().send_document(xlsx, caption=caption)
        ok = bool(getattr(res, "success", False))
        logger.info("Telegram: daily Excel for {} {}.", eval_date, "sent" if ok else "send failed")
        return ok
    except Exception as exc:  # noqa: BLE001
        logger.warning("Telegram: daily Excel send skipped ({}).", exc)
        return False


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Signal List performance tracker")
    p.add_argument("--date", default=None, help="Only this signal date (YYYY-MM-DD).")
    p.add_argument("--evening", action="store_true",
                   help="Evening mode: fetch same-day OHLC for pending tickers, then evaluate.")
    p.add_argument("--telegram", action="store_true",
                   help="After evaluating, send the latest daily Excel to Telegram (fail-safe).")
    a = p.parse_args()
    if a.evening:
        run_performance_evening()
    else:
        run_performance(dates=[a.date] if a.date else None)
    if a.telegram:
        send_daily_excel_telegram()
